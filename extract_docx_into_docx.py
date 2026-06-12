"""
Create a copy of a .docx file with all embedded OLE objects (xlsx, xlsm, csv,
msg, eml, txt, docx, pdf) replaced by their extracted text content at the same
positions within the document.

Usage:
    python3 extract_docx_separate.py <path_to_docx> [output_path]

If output_path is not specified, the output is saved as <input_name>_extracted.docx
"""

import zipfile
import os
import struct
import email
import copy
from xml.etree import ElementTree as ET
from io import BytesIO

import openpyxl
import olefile


# XML namespaces used in .docx files
NSMAP = {
    'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'o': 'urn:schemas-microsoft-com:office:office',
    'v': 'urn:schemas-microsoft-com:vml',
    'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
    'wp': 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'w14': 'http://schemas.microsoft.com/office/word/2010/wordml',
}

for prefix, uri in NSMAP.items():
    ET.register_namespace(prefix, uri)

PROGID_EXT_MAP = {
    'Word.Document.12': '.docx',
    'Word.Document.8': '.doc',
    'Excel.Sheet.12': '.xlsx',
    'Excel.Sheet.8': '.xls',
    'Excel.SheetMacroEnabled.12': '.xlsm',
    'Acrobat.Document.DC': '.pdf',
    'Acrobat.Document': '.pdf',
    'Package': '',
}


def extract_xlsx_text(data: bytes) -> list[str]:
    wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        lines.append(f"[Sheet: {sheet}]")
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_text = '\t'.join(str(c) if c is not None else '' for c in row)
            if row_text.strip():
                lines.append(row_text)
    wb.close()
    return lines


def extract_csv_text(data: bytes) -> list[str]:
    try:
        text = data.decode('utf-8')
    except UnicodeDecodeError:
        text = data.decode('latin-1')
    return [line for line in text.splitlines() if line.strip()]


def extract_txt_text(data: bytes) -> list[str]:
    try:
        text = data.decode('utf-8')
    except UnicodeDecodeError:
        text = data.decode('latin-1')
    return [line for line in text.splitlines() if line.strip()]


def extract_eml_text(data: bytes) -> list[str]:
    try:
        text = data.decode('utf-8')
    except UnicodeDecodeError:
        text = data.decode('latin-1')
    try:
        msg = email.message_from_string(text)
        parts = []
        if msg['Subject']:
            parts.append(f"Subject: {msg['Subject']}")
        if msg['From']:
            parts.append(f"From: {msg['From']}")
        if msg['To']:
            parts.append(f"To: {msg['To']}")
        if msg['Date']:
            parts.append(f"Date: {msg['Date']}")
        body = msg.get_payload(decode=True)
        if body:
            parts.append(f"Body: {body.decode('utf-8', errors='replace')}")
        elif isinstance(msg.get_payload(), str):
            parts.append(f"Body: {msg.get_payload()}")
        return parts
    except Exception:
        return [text]


def extract_msg_text(data: bytes) -> list[str]:
    try:
        import extract_msg
        msg = extract_msg.openMsg(BytesIO(data))
        parts = []
        if msg.subject:
            parts.append(f"Subject: {msg.subject}")
        if msg.sender:
            parts.append(f"From: {msg.sender}")
        if msg.to:
            parts.append(f"To: {msg.to}")
        if msg.body:
            parts.append(f"Body: {msg.body}")
        return parts
    except Exception:
        try:
            text = data.decode('utf-8')
        except UnicodeDecodeError:
            text = data.decode('latin-1')
        return [text]


def extract_pdf_text(data: bytes) -> list[str]:
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(BytesIO(data))
        pages = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text()
            if text and text.strip():
                pages.append(f"[Page {i+1}]")
                pages.append(text.strip())
        return pages if pages else ["[Empty PDF]"]
    except ImportError:
        try:
            import pdfplumber
            with pdfplumber.open(BytesIO(data)) as pdf:
                pages = []
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if text and text.strip():
                        pages.append(f"[Page {i+1}]")
                        pages.append(text.strip())
                return pages if pages else ["[Empty PDF]"]
        except ImportError:
            return ["[PDF detected but PyPDF2/pdfplumber not installed]"]


def extract_embedded_docx_text(data: bytes) -> list[str]:
    import docx
    doc = docx.Document(BytesIO(data))
    return [p.text for p in doc.paragraphs if p.text.strip()]


def extract_ole_embedded(data: bytes, prog_id: str = ''):
    try:
        ole = olefile.OleFileIO(BytesIO(data))
        if ole.exists('CONTENTS'):
            stream = ole.openstream('CONTENTS')
            file_data = stream.read()
            ole.close()
            ext = PROGID_EXT_MAP.get(prog_id, '')
            filename = f'embedded{ext}' if ext else 'embedded_contents'
            return filename, file_data
        if ole.exists('\x01Ole10Native'):
            stream = ole.openstream('\x01Ole10Native')
            content = stream.read()
            ole.close()
            return parse_ole10native(content)
        if ole.exists('Package'):
            stream = ole.openstream('Package')
            file_data = stream.read()
            ole.close()
            return 'unknown_package', file_data
        ole.close()
    except Exception:
        pass
    return None, None


def parse_ole10native(content: bytes):
    if len(content) < 8:
        return None, None
    idx = 0
    total_size = struct.unpack('<I', content[idx:idx+4])[0]
    idx += 4
    flags = struct.unpack('<H', content[idx:idx+2])[0]
    idx += 2
    end = content.index(b'\x00', idx)
    label = content[idx:end].decode('latin-1')
    idx = end + 1
    end = content.index(b'\x00', idx)
    src_path = content[idx:end].decode('latin-1')
    idx = end + 1
    filename = label if label else os.path.basename(src_path)
    if idx + 4 <= len(content):
        next_dword = struct.unpack('<I', content[idx:idx+4])[0]
        if next_dword == 0x00030000:
            idx += 4
            temp_len = struct.unpack('<I', content[idx:idx+4])[0]
            idx += 4
            idx += temp_len
            if idx + 4 <= len(content):
                data_size = struct.unpack('<I', content[idx:idx+4])[0]
                idx += 4
                file_data = content[idx:idx+data_size]
                return filename, file_data
        else:
            try:
                end = content.index(b'\x00', idx)
                idx = end + 1
                end = content.index(b'\x00', idx)
                idx = end + 1
                if idx + 4 <= len(content):
                    data_size = struct.unpack('<I', content[idx:idx+4])[0]
                    idx += 4
                    file_data = content[idx:idx+data_size]
                    return filename, file_data
            except ValueError:
                pass
    return None, None


def guess_extension_from_data(data: bytes) -> str:
    if data[:4] == b'PK\x03\x04':
        try:
            zf = zipfile.ZipFile(BytesIO(data))
            names = zf.namelist()
            if any('word/' in n for n in names):
                return '.docx'
            elif any('xl/' in n for n in names):
                return '.xlsx'
            elif any('ppt/' in n for n in names):
                return '.pptx'
            zf.close()
        except Exception:
            pass
        return '.zip'
    elif data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return '.ole'
    elif data[:5] == b'%PDF-':
        return '.pdf'
    elif data[:5] == b'MIME-' or data[:9] == b'Received:' or data[:5] == b'From:':
        return '.eml'
    return ''


def read_embedded_file(filename: str, data: bytes) -> list[str]:
    ext = os.path.splitext(filename)[1].lower()
    if not ext:
        ext = guess_extension_from_data(data)
    if ext in ('.xlsx', '.xlsm'):
        return extract_xlsx_text(data)
    elif ext == '.csv':
        return extract_csv_text(data)
    elif ext == '.txt':
        return extract_txt_text(data)
    elif ext == '.eml':
        return extract_eml_text(data)
    elif ext == '.msg':
        return extract_msg_text(data)
    elif ext == '.docx':
        return extract_embedded_docx_text(data)
    elif ext == '.pdf':
        return extract_pdf_text(data)
    else:
        return [f"[Unsupported format: {ext or 'unknown'}]"]


def make_text_paragraph(text: str, para_props_elem=None) -> ET.Element:
    """Create a w:p element containing a text run."""
    w = NSMAP['w']
    p = ET.Element(f'{{{w}}}p')
    if para_props_elem is not None:
        p.append(copy.deepcopy(para_props_elem))
    r = ET.SubElement(p, f'{{{w}}}r')
    t = ET.SubElement(r, f'{{{w}}}t')
    t.set(f'{{{NSMAP["w"]}}}space', 'preserve')
    t.text = text
    return p


def make_separator_paragraph(text: str, para_props_elem=None) -> ET.Element:
    """Create a bold separator paragraph (for --- Embedded File: ... --- markers)."""
    w = NSMAP['w']
    p = ET.Element(f'{{{w}}}p')
    if para_props_elem is not None:
        p.append(copy.deepcopy(para_props_elem))
    r = ET.SubElement(p, f'{{{w}}}r')
    rpr = ET.SubElement(r, f'{{{w}}}rPr')
    ET.SubElement(rpr, f'{{{w}}}b')
    ET.SubElement(rpr, f'{{{w}}}sz').set(f'{{{w}}}val', '18')
    t = ET.SubElement(r, f'{{{w}}}t')
    t.set(f'{{{NSMAP["w"]}}}space', 'preserve')
    t.text = text
    return p


def resolve_ole_object(obj, rels, rel_types, zf):
    """Given an OLEObject element, resolve it to (filename, file_data) or (None, None)."""
    rid = obj.get(f'{{{NSMAP["r"]}}}id')
    if not rid or rid not in rels:
        return None, None

    target = rels[rid]
    rel_type = rel_types.get(rid, '')
    prog_id = obj.get('ProgID', '')
    embed_path = f'word/{target}' if not target.startswith('/') else target.lstrip('/')

    try:
        embed_data = zf.read(embed_path)
    except KeyError:
        return None, None

    filename = os.path.basename(target)
    file_data = None

    if 'package' in rel_type.lower():
        file_data = embed_data
        ext = os.path.splitext(filename)[1].lower()
        if not ext or ext == '.bin':
            prog_ext = PROGID_EXT_MAP.get(prog_id, '')
            if prog_ext:
                filename = os.path.splitext(filename)[0] + prog_ext
    elif 'oleobject' in rel_type.lower().replace('/', ''):
        ole_filename, ole_data = extract_ole_embedded(embed_data, prog_id)
        if ole_filename and ole_data:
            filename = ole_filename
            file_data = ole_data
        else:
            file_data = embed_data
    else:
        ole_filename, ole_data = extract_ole_embedded(embed_data, prog_id)
        if ole_filename and ole_data:
            filename = ole_filename
            file_data = ole_data
        else:
            file_data = embed_data

    return filename, file_data


def process_paragraph_for_replacement(para, rels, rel_types, zf):
    """
    Process a paragraph element. If it contains OLE objects, return replacement
    paragraph elements with extracted text. Otherwise return [para] unchanged.
    """
    w = NSMAP['w']
    o = NSMAP['o']

    ole_objects = list(para.iter(f'{{{o}}}OLEObject'))
    if not ole_objects:
        return [para]

    # Get paragraph properties to replicate styling
    para_props = para.find(f'{{{w}}}pPr')

    # Collect text from this paragraph (non-OLE runs)
    para_text = ''.join(
        node.text or ''
        for node in para.iter(f'{{{w}}}t')
    )

    replacements = []

    for obj in ole_objects:
        filename, file_data = resolve_ole_object(obj, rels, rel_types, zf)
        if not filename or not file_data:
            continue

        # Add preceding paragraph text if any
        if para_text.strip():
            replacements.append(make_text_paragraph(para_text, para_props))
            para_text = ''

        # Add separator and extracted content
        replacements.append(make_separator_paragraph(
            f"--- Embedded File: {filename} ---", para_props))

        lines = read_embedded_file(filename, file_data)
        for line in lines:
            replacements.append(make_text_paragraph(line, para_props))

        replacements.append(make_separator_paragraph(
            f"--- End: {filename} ---", para_props))

    # If we produced replacements, return them; otherwise keep original
    if replacements:
        # If there's remaining text after OLE (unlikely but possible)
        if para_text.strip():
            replacements.append(make_text_paragraph(para_text, para_props))
        return replacements

    return [para]


def process_document(docx_path: str, output_path: str):
    """Create a copy of the docx with OLE objects replaced by extracted text."""
    zf = zipfile.ZipFile(docx_path)

    # Parse relationships
    rels_xml = zf.read('word/_rels/document.xml.rels')
    rels_tree = ET.fromstring(rels_xml)
    rels = {}
    rel_types = {}
    for rel in rels_tree:
        rid = rel.get('{http://schemas.openxmlformats.org/package/2006/relationships}Id')
        if rid is None:
            rid = rel.get('Id')
        target = rel.get('Target')
        rel_type = rel.get('Type', '')
        if rid and target:
            rels[rid] = target
            rel_types[rid] = rel_type

    # Parse document XML
    doc_xml = zf.read('word/document.xml')
    tree = ET.fromstring(doc_xml)
    body = tree.find(f'.//{{{NSMAP["w"]}}}body')

    w = NSMAP['w']
    o = NSMAP['o']

    # Process body elements - replace OLE-containing paragraphs with extracted text
    new_body_children = []
    for element in list(body):
        tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag

        if tag == 'p':
            new_body_children.extend(
                process_paragraph_for_replacement(element, rels, rel_types, zf))

        elif tag == 'tbl':
            # Process table: walk cells and replace OLE paragraphs within them
            for cell in element.iter(f'{{{w}}}tc'):
                # Collect current children of the cell
                cell_children = list(cell)
                # Check if any paragraph in this cell has OLE
                has_ole = any(
                    p.find(f'.//{{{o}}}OLEObject') is not None
                    for p in cell.iter(f'{{{w}}}p')
                )
                if not has_ole:
                    continue

                # Rebuild cell contents
                new_cell_children = []
                for child in cell_children:
                    child_tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if child_tag == 'p':
                        new_cell_children.extend(
                            process_paragraph_for_replacement(child, rels, rel_types, zf))
                    else:
                        new_cell_children.append(child)

                # Replace cell contents
                for child in list(cell):
                    cell.remove(child)
                for child in new_cell_children:
                    cell.append(child)

            new_body_children.append(element)
        else:
            new_body_children.append(element)

    # Replace body contents
    for child in list(body):
        body.remove(child)
    for child in new_body_children:
        body.append(child)

    # Serialize modified document XML
    modified_doc_xml = ET.tostring(tree, encoding='unicode', xml_declaration=True)

    # Write the output docx: copy all files, replacing document.xml
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as out_zf:
        for item in zf.infolist():
            if item.filename == 'word/document.xml':
                out_zf.writestr(item, modified_doc_xml)
            else:
                out_zf.writestr(item, zf.read(item.filename))

    zf.close()
    print(f"Output saved to: {output_path}")


if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("Usage: python3 extract_docx_separate.py <path_to_docx> [output_path]")
        sys.exit(1)

    input_path = sys.argv[1]
    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_extracted{ext}"

    process_document(input_path, output_path)
